from io import BytesIO
from datetime import date
import os

from flask import Blueprint, render_template, request, send_file, current_app, flash, redirect, url_for
from flask_login import login_required

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.units import cm

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from utils.decorators import role_required
from utils.db import query_one, query_all


bp = Blueprint("laporan", __name__, template_folder="../../templates")


def _get_range_from_args():
    start = request.args.get("start") or date.today().isoformat()
    end = request.args.get("end") or date.today().isoformat()
    return start, end


def _fetch_absensi(start: str, end: str):
    rows = query_all(
        """
        SELECT a.tanggal, a.jam_masuk, a.jam_pulang, a.status, a.terlambat, a.keterangan, a.validated,
               g.nip, g.nama, g.jabatan, g.mata_pelajaran
        FROM absensi a
        JOIN guru g ON g.id = a.guru_id
        WHERE a.tanggal BETWEEN ? AND ?
        ORDER BY a.tanggal ASC, g.nama ASC
        """,
        (start, end),
    )
    return rows


@bp.route("/")
@login_required
@role_required("admin")
def index():
    start, end = _get_range_from_args()
    return render_template("laporan/index.html", start=start, end=end)


@bp.route("/pdf")
@login_required
@role_required("admin")
def export_pdf():
    start, end = _get_range_from_args()
    prof = query_one("SELECT * FROM profil_madrasah WHERE id=1")
    rows = _fetch_absensi(start, end)

    buff = BytesIO()
    doc = SimpleDocTemplate(buff, pagesize=A4, leftMargin=2 * cm, rightMargin=2 * cm)
    styles = getSampleStyleSheet()
    story = []

    # header with logos
    logo_left = None
    logo_right = None
    try:
        if prof and prof["logo_kemenag"]:
            logo_right = Image(
                os.path.join(current_app.static_folder, prof["logo_kemenag"]),
                width=2.4 * cm,
                height=2.4 * cm,
            )
        if prof and prof["logo_madrasah"]:
            logo_left = Image(
                os.path.join(current_app.static_folder, prof["logo_madrasah"]),
                width=2.4 * cm,
                height=2.4 * cm,
            )
    except Exception:
        logo_left = None
        logo_right = None

    title = Paragraph(
        f"<b>LAPORAN ABSENSI GURU</b><br/>{prof['nama_madrasah'] if prof else ''}",
        styles["Title"],
    )
    sub = Paragraph(
        f"Periode: <b>{start}</b> s/d <b>{end}</b><br/>{prof['alamat'] if prof else ''}",
        styles["Normal"],
    )

    header_data = [[logo_right or "", title, logo_left or ""]]
    header_tbl = Table(header_data, colWidths=[3 * cm, 11 * cm, 3 * cm])
    header_tbl.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (0, 0), "LEFT"),
                ("ALIGN", (2, 0), (2, 0), "RIGHT"),
                ("ALIGN", (1, 0), (1, 0), "CENTER"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(header_tbl)
    story.append(sub)
    story.append(Spacer(1, 12))

    table_data = [
        [
            "No",
            "Tanggal",
            "NIP",
            "Nama",
            "Status",
            "Masuk",
            "Pulang",
            "Terlambat (mnt)",
            "Keterangan",
            "Valid",
        ]
    ]
    for i, r in enumerate(rows, 1):
        table_data.append(
            [
                i,
                r["tanggal"],
                r["nip"],
                r["nama"],
                r["status"],
                r["jam_masuk"] or "-",
                r["jam_pulang"] or "-",
                r["terlambat"] or 0,
                r["keterangan"] or "",
                "Ya" if (r["validated"] or 0) == 1 else "Tidak",
            ]
        )

    tbl = Table(table_data, repeatRows=1)
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F9D58")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
            ]
        )
    )
    story.append(tbl)
    story.append(Spacer(1, 18))

    # signature block
    kepala = prof["kepala_madrasah"] if prof else ""
    ttd_path = None
    if prof and prof["ttd_kepala"]:
        candidate = os.path.join(current_app.static_folder, prof["ttd_kepala"])
        if os.path.exists(candidate):
            ttd_path = candidate

    sign_tbl_data = [["Mengetahui,", "", f"{date.today().isoformat()}"], [f"Kepala Madrasah", "", ""]]
    sign_tbl = Table(sign_tbl_data, colWidths=[6 * cm, 4 * cm, 6 * cm])
    sign_tbl.setStyle(TableStyle([("ALIGN", (0, 0), (-1, -1), "CENTER")]))
    story.append(sign_tbl)
    story.append(Spacer(1, 12))
    if ttd_path:
        try:
            story.append(Image(ttd_path, width=4 * cm, height=2 * cm))
        except Exception:
            pass
    story.append(Spacer(1, 6))
    story.append(Paragraph(f"<b><u>{kepala}</u></b>", styles["Normal"]))

    doc.build(story)
    buff.seek(0)

    filename = f"laporan_absensi_{start}_sd_{end}.pdf"
    return send_file(buff, as_attachment=True, download_name=filename, mimetype="application/pdf")


@bp.route("/excel")
@login_required
@role_required("admin")
def export_excel():
    start, end = _get_range_from_args()
    prof = query_one("SELECT * FROM profil_madrasah WHERE id=1")
    rows = _fetch_absensi(start, end)

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Absensi"

    header_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="0F9D58")

    ws["A1"] = "LAPORAN ABSENSI GURU"
    ws["A2"] = prof["nama_madrasah"] if prof else ""
    ws["A3"] = f"Periode: {start} s/d {end}"
    for cell in ["A1", "A2", "A3"]:
        ws[cell].font = Font(bold=True)

    cols = ["No", "Tanggal", "NIP", "Nama", "Status", "Masuk", "Pulang", "Terlambat", "Keterangan", "Valid"]
    start_row = 5
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=start_row, column=j, value=c)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, r in enumerate(rows, 1):
        rr = start_row + i
        ws.cell(rr, 1, i)
        ws.cell(rr, 2, r["tanggal"])
        ws.cell(rr, 3, r["nip"])
        ws.cell(rr, 4, r["nama"])
        ws.cell(rr, 5, r["status"])
        ws.cell(rr, 6, r["jam_masuk"] or "-")
        ws.cell(rr, 7, r["jam_pulang"] or "-")
        ws.cell(rr, 8, int(r["terlambat"] or 0))
        ws.cell(rr, 9, r["keterangan"] or "")
        ws.cell(rr, 10, "Ya" if (r["validated"] or 0) == 1 else "Tidak")

    # autosize
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"laporan_absensi_{start}_sd_{end}.xlsx"
    return send_file(
        out,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
